import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from io import BytesIO
import time
import os
import webbrowser
import requests
import tempfile

# Function to get email contacts from People column
def get_emails(people):
    if pd.isna(people):
        return []
    contacts = people.split(';')
    emails = [contact.split('<')[1].replace('>', '').strip() for contact in contacts]
    return emails[:3]  # Only consider the first three contacts

# Function to format dates to a more user-friendly format
def format_date(date_str):
    if pd.isna(date_str):
        return ""
    date_obj = datetime.fromisoformat(date_str)
    return date_obj.strftime('%Y-%m-%d %H:%M')

# Function to extract relevant data and populate the Excel file
def populate_excel(ws, export_df, notes_df, persons_df):
    start_row = 22

    # Add data from export_df
    for index, row in export_df.iterrows():
        acquirer_name = row['Name'].split(' - ')[1]
        ws.cell(row=start_row + index, column=1, value=row['Wave/Tier'])  # Wave
        ws.cell(row=start_row + index, column=2, value=acquirer_name)  # Acquirer's Name
        ws.cell(row=start_row + index, column=4, value=row['Buyer Status'])  # Status
        ws.cell(row=start_row + index, column=5, value=row['Introduction Call'])  # Introduction Call
        ws.cell(row=start_row + index, column=6, value=row['Management Presentation'])  # Management Presentation
        ws.cell(row=start_row + index, column=7, value=row['NDA Signed'])  # NDA Signed

        # Extract and add contacts data
        emails = get_emails(row['People'])
        for i, email in enumerate(emails):
            # Ensure there are no NaN values before filtering
            person_data = persons_df[persons_df['Emails'].fillna('').str.contains(email)]
            if not person_data.empty:
                person = person_data.iloc[0]
                ws.cell(row=start_row + index, column=9 + i * 4, value=person['Full Name'])  # Surname / name contact
                ws.cell(row=start_row + index, column=10 + i * 4, value=person['Job Titles'])  # Position
                ws.cell(row=start_row + index, column=11 + i * 4, value=person['Emails'])  # Email
                ws.cell(row=start_row + index, column=12 + i * 4, value=person['LinkedIn Url'])  # LinkedIn Url

    # Add data from notes_df
    for index, row in notes_df.iterrows():
        opportunity_name = row['Opportunity'].split(' - ')[1]
        formatted_date = format_date(row['Author Date'])
        for cell in ws['B']:  # Searching in the Acquirer's Name column
            if cell.value == opportunity_name:
                row_num = cell.row
                ws.cell(row=row_num, column=21, value=row['Content'])  # Notes
                ws.cell(row=row_num, column=22, value=formatted_date)  # Notes date
                break

# Function to determine the type of CSV file
def determine_file_type(file):
    df = pd.read_csv(file)
    if 'Wave/Tier' in df.columns:
        return 'export'
    elif 'Author Date' in df.columns:
        return 'notes'
    elif 'Emails' in df.columns:
        return 'persons'
    else:
        return None

# Streamlit app
st.set_page_config(page_title="Roadshow Data Populator", page_icon="images/IPTP.png", layout="centered", initial_sidebar_state="expanded")

# Custom CSS for styling
st.markdown("""
    <style>
        .header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 10px;
        }
        .header img {
            width: 150px;
        }
        .header h1 {
            flex-grow: 1;
            text-align: center;
            margin: 0;
        }
        .sidebar-footer {
            position: fixed;
            bottom: 0;
            width: 100%;
            text-align: center;
        }
        .success-upload {
            color: green;
            font-weight: bold;
        }
        .styled-button {
            border: 2px solid #4CAF50;
            color: #4CAF50;
            padding: 10px 20px;
            text-align: center;
            text-decoration: none;
            display: inline-block;
            font-size: 16px;
            margin: 4px 2px;
            cursor: pointer;
            border-radius: 12px;
            background-color: white;
        }
        .styled-button-view {
            background-color: #008CBA; /* Blue */
            border: none;
            color: white;
            padding: 15px 32px;
            text-align: center;
            text-decoration: none;
            display: inline-block;
            font-size: 16px;
            margin: 4px 2px;
            cursor: pointer;
            border-radius: 12px;
        }
        .center-button {
            display: flex;
            justify-content: center;
        }
        .file-upload {
            border: 2px dashed #ccc;
            border-radius: 10px;
            padding: 20px;
            text-align: center;
        }
        .button-row {
            display: flex;
            justify-content: space-between;
        }
    </style>
""", unsafe_allow_html=True)

# Header with logos
col1, col2, col3 = st.columns([1, 3, 1])
with col1:
    st.image("images/IPTP.png", width=80)
with col2:
    st.markdown("<h1 style='text-align: center;'>Roadshow Data Populator</h1>", unsafe_allow_html=True)
with col3:
    st.image("images/Affinity.svg", width=200)

st.sidebar.title("About")
st.sidebar.info("""
This app allows you to upload Project Tracker CSV files and populates an Excel template with the data.

There is an option to drop all csv files at once instead of uploading them separately

""")

# Option to drop all files or upload individually
file_upload_option = st.sidebar.radio("File Upload Option", ("Drop All Files", "Upload Individually"))
if file_upload_option == "Upload Individually":
    st.info(
        """
        Upload your CSV files to populate the Roadshow Excel template.
        """,
        icon="📄",
    )

if file_upload_option == "Drop All Files":
    st.info(
        """
        Drop all CSV files here to populate the Roadshow Excel template.
        """,
        icon="📁",
    )

if file_upload_option == "Upload Individually":
    # Upload CSV files individually
    export_file = st.file_uploader("Upload Export CSV", type="csv", help="Upload the export CSV file", key="export_csv")
    if export_file:
        st.markdown('<p class="success-upload">File Uploaded Successfully!</p>', unsafe_allow_html=True)

    notes_file = st.file_uploader("Upload Notes CSV (commentaires)", type="csv", help="Upload the notes CSV file", key="notes_csv")
    if notes_file:
        st.markdown('<p class="success-upload">File Uploaded Successfully!</p>', unsafe_allow_html=True)

    persons_file = st.file_uploader("Upload Associated Persons CSV (Contacts data)", type="csv", help="Upload the associated persons CSV file", key="persons_csv")
    if persons_file:
        st.markdown('<p class="success-upload">File Uploaded Successfully!</p>', unsafe_allow_html=True)
else:
    # Upload all CSV files at once
    all_files = st.file_uploader("Drop all CSV files here", type="csv", accept_multiple_files=True, help="Upload all CSV files together", key="all_files")
    export_file, notes_file, persons_file = None, None, None
    if all_files:
        for file in all_files:
            file_type = determine_file_type(file)
            file.seek(0)  # Reset file pointer after reading
            if file_type == 'export':
                export_file = file
            elif file_type == 'notes':
                notes_file = file
            elif file_type == 'persons':
                persons_file = file

        # Check for missing files
        if not export_file:
            st.error("Export CSV file is missing. Please upload it again.")
        if not notes_file:
            st.error("Notes CSV file is missing. Please upload it again.")
        if not persons_file:
            st.error("Persons CSV file is missing. Please upload it again.")

# Modify Template button to toggle visibility of the new template uploader
if 'show_new_template_uploader' not in st.session_state:
    st.session_state.show_new_template_uploader = False

# Button row for View Template and Modify Template
col1, col2 = st.columns(2)
with col1:
    st.button('Modify Template', on_click=lambda: setattr(st.session_state, 'show_new_template_uploader', not st.session_state.show_new_template_uploader))
    if st.session_state.show_new_template_uploader:
        new_template_file = st.file_uploader("Upload a New Excel Template (optional)", type="xlsx", help="Upload a new Excel template for this session", key="new_template")
        if new_template_file:
            st.markdown('<p class="success-upload">Template Uploaded Successfully!</p>', unsafe_allow_html=True)
            template_path = new_template_file
        else:
            template_path = 'templates/Roadshow_template.xlsx'
    else:
        template_path = 'templates/Roadshow_template.xlsx'

with col2:
    if 'template_downloaded' not in st.session_state:
        st.session_state.template_downloaded = False

    def download_and_open_template():
        with open('templates/Roadshow_template.xlsx', 'rb') as f:
            st.download_button(
                label='Download Template',
                data=f,
                file_name='Roadshow_template.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        st.session_state.template_downloaded = True

    if st.button('View Original Excel Template'):
        download_and_open_template()


# Show progress bar only for generating the Excel file
if export_file and notes_file and persons_file:
    # Read CSV files
    export_df = pd.read_csv(export_file)
    notes_df = pd.read_csv(notes_file)
    persons_df = pd.read_csv(persons_file)

    # Load the template Excel file
    wb = load_workbook(template_path)
    ws = wb['Suivi du Roadshow']  # Assuming the sheet to be filled is named 'Suivi du Roadshow'

    # Get the dossier name from the export CSV
    dossier_name = export_df.iloc[0]['Name'].split(' - ')[0]
    ws['A1'] = f"{dossier_name} - Roadshow"

    # Add the current month and year below the title
    current_date = datetime.now()
    month_year = current_date.strftime('%B %Y')
    ws['A2'] = month_year

    # Show progress bar
    if 'progress_bar_visible' not in st.session_state:
        st.session_state.progress_bar_visible = True

    if st.session_state.progress_bar_visible:
        progress_bar = st.progress(0)
        progress_text = st.empty()
        for i in range(100):
            time.sleep(0.01)
            progress_bar.progress(i + 1)
            progress_text.text(f"Generating Excel file: {i + 1}%")

    # Populate the Excel file
    populate_excel(ws, export_df, notes_df, persons_df)

    # Save the updated Excel file to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("Data has been successfully populated into the Excel template.")

    st.markdown("### Generated Excel File Preview")
    # Make row 19 the header and rows below as data
    updated_df = pd.read_excel(output, sheet_name='Suivi du Roadshow', header=20)
    updated_df = updated_df[['Wave', "Acquirer's Name", 'Status', 'Intro call', 'Tech call', 'NDA signed', 'Surname / Name contact 1', 'Position contact 1', 'Contact shooté 1', 'LinkedIn contact 1', 'Surname / Name contact 2', 'Position contact 2', 'Contact shooté 2', 'LinkedIn contact 2', 'Surname / Name contact 3', 'Position contact 3', 'Contact shooté 3', 'LinkedIn contact 3', 'Comments / Rationale (if passed)', 'Date of comments']]
    st.dataframe(updated_df)

    # Display buttons for viewing and downloading the generated file
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="Download Generated Excel File",
            data=output,
            file_name=f'Generated_Roadshow_{dossier_name}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    st.image("Gifs/arrow_small_new.gif")

    # If permission error, display the dataframe as a fallback option
    if st.session_state.get('permission_error', False):
        st.error("Permission denied: Unable to save or access the file. Please ensure the file is not open in another program and try again.")
        st.info("You can view the generated file below:")
        updated_df = pd.read_excel(output, sheet_name='Suivi du Roadshow', header=20)
        updated_df = updated_df[['Wave', "Acquirer's Name", 'Status', 'Intro call', 'Tech call', 'NDA signed', 'Surname / Name contact 1', 'Position contact 1', 'Contact shooté 1', 'LinkedIn contact 1', 'Surname / Name contact 2', 'Position contact 2', 'Contact shooté 2', 'LinkedIn contact 2', 'Surname / Name contact 3', 'Position contact 3', 'Contact shooté 3', 'LinkedIn contact 3', 'Comments / Rationale (if passed)', 'Date of comments']]
        st.dataframe(updated_df)
        
st.sidebar.markdown("""
    <style>
        .custom-text {
            font-family: 'Arial', sans-serif;
            font-size: 16px;
            color: #000000;
        }
        .custom-text strong {
            font-weight: bold;
            color: #000000;
        }
        .github-link {
            text-align: center;
            margin-top: 100px;
            position: absolute;
            bottom: -280px;
            width: 100%;
        }
        .github-link a {
            text-decoration: underline;
            color: #0000EE;
        }
        .github-link img {
            width: 25px;
            height: 25px;
            vertical-align: middle;
        }
    </style>

    <div class="github-link">
        <img src="https://logos-download.com/wp-content/uploads/2016/09/GitHub_logo.png" alt="GitHub"> 
        Developed by <a href="https://github.com/Ayman2G" target="_blank">@Ayman2G</a>
    </div>
""", unsafe_allow_html=True)
